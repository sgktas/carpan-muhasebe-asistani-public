from pathlib import Path

from openpyxl import Workbook

from app.core.tahsilat_parser import TahsilatParser


HEADERS = ["MusteriKodu", "Musteriİsmi", "BelgeTarihi", "Tutar"]


def _append_sheet(workbook: Workbook, title: str, rows: list[list[object]]):
    worksheet = workbook.create_sheet(title)
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append(row)
    return worksheet


def test_single_sheet_tahsilat_raporu_ilk_sayfadan_okunur(tmp_path: Path):
    path = tmp_path / "tahsilat.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Tahsilat"
    worksheet.append(HEADERS)
    worksheet.append(["C001", "TEST MUSTERI", "21.07.2026", 100.0])
    workbook.save(path)

    parser = TahsilatParser(path)
    records = parser.load()

    assert parser.selected_sheet_name == 0
    assert len(records) == 1
    assert records[0].musteri_kodu == "C001"


def test_rapor_duzenleme_ciktisinda_subeliler_sayfasi_tercih_edilir(tmp_path: Path):
    path = tmp_path / "duzenlenmis_tahsilat.xlsx"
    workbook = Workbook()
    main = workbook.active
    main.title = "ENT-Muhasebe Entegrasyon"
    main.append(HEADERS)
    main.append(["C001", "ANA SAYFA KAYDI", "21.07.2026", 100.0])

    _append_sheet(
        workbook,
        "ŞUBELİLER",
        [
            ["C001", "ANA SAYFA KAYDI", "21.07.2026", 100.0],
            ["C002", "SUBELI MUSTERI", "21.07.2026", 250.0],
        ],
    )
    workbook.save(path)

    parser = TahsilatParser(path)
    records = parser.load()

    assert parser.selected_sheet_name == "ŞUBELİLER"
    assert len(records) == 2
    assert {record.musteri_kodu for record in records} == {"C001", "C002"}
    assert sum(record.tutar for record in records) == 350.0
