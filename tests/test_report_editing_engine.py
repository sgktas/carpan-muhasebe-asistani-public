from datetime import datetime
from pathlib import Path

import xlrd
import xlwt
from openpyxl import Workbook, load_workbook

from app.modules.report_editing.engine import (
    COLLECTION_CLEAN_OUTPUT_PREFIX,
    COLLECTION_OUTPUT_BASENAME,
    COLLECTION_SHEET_NAME,
    SALES_CLEAN_OUTPUT_PREFIX,
    SALES_OUTPUT_BASENAME,
    SALES_SHEET_NAME,
    ExcelTemplateWriter,
    ReportEditingEngine,
    refresh_customer_list_cache,
)
from app.core.customer_list_cache import CustomerListCache


def _save_xls(path: Path, sheet_name: str, headers: list[str], rows: list[list[object]]) -> Path:
    """xlwt ile eski (BIFF/.xls) formatta test dosyası üretir — FOM'un yeni
    dışa aktarım biçimini taklit eder."""
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet(sheet_name[:31])
    date_style = xlwt.XFStyle()
    date_style.num_format_str = "DD.MM.YYYY"
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)
    for row_index, row in enumerate(rows, start=1):
        for col, value in enumerate(row):
            if isinstance(value, datetime):
                worksheet.write(row_index, col, value, date_style)
            else:
                worksheet.write(row_index, col, value)
    workbook.save(str(path))
    return path


def _sales_headers() -> list[str]:
    return [
        "MüşteriKodu", "FaturaNo", "Tarih", "KDV", "PersonelKodu", "ÖdemeTipi",
        "ÜrünKodu", "FOC", "Tabela Adı", "Vergi Dairesi", "Vergi No",
        "İlk Matbu No", "Fatura Kodu", "İrsaliye Kodu", "İrsaliye Numarası",
        "İrsaliye Tarihi", "Miktar", "Fiyat", "İskonto1", "İskonto2",
        "ToplamKDV", "EklenenKDV", "Vade", "TuketiciFiyati", "NetFiyat",
    ]


def _save(path: Path, headers: list[str], rows: list[list[object]]) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    return path


def _customer_headers() -> list[str]:
    return [
        "Şube", "Müşteri Sayısı", "Müşteri Kodu", "Tabela Adi", "Ünvan",
        "Vergi Dairesi", "Vergi Numarası", "Tekel No", "Aidiyet", "eMail",
        "Satış Kapama Nedeni", "Durum Nedeni", "Aktif/Pasif Nedeni",
        "Yeni Müşteri Kodu", "Eski Müşteri Kodu", "Durum", "İlgili Kişi",
        "Adres", "İl", "İlçe", "Bakiye", "Müşteri Kayıt Tipi", "Müşteri Tipi",
        "Alt Tip", "Sahiplik Tipi", "SR-Rota", "Banka", "Ödeme Tipi",
        "Faks No", "Risk Limiti", "Etiket", "Telefon", "Kredi Limiti",
        "Fiyat Listesi", "ERP Kodu", "EFatura", "E-fat-Tip", "Vergi Tipi",
        "Hero Sınıfı", "Dış Kay.Sip.Kod",
    ]


def test_report_engine_applies_reference_rules(tmp_path):
    customer = _save(
        tmp_path / "customer.xlsx",
        _customer_headers(),
        [
            [
                "SIMSEK-AYDIN", 1, "1001", "TEST MARKET", "TEST LTD", "AYDIN",
                "1111111111", None, None, None, None, None, None, None, None,
                "Aktif", None, None, "AYDIN", "NAZILLI", 0, "Müşteri", "Müşteri",
                "Market", "Bagimsiz Satis Noktasi",
                "PERSONEL - AYDIN-DD-0203", None, "Açık Hesap", None, 0, None,
                "555", 0, "Liste", "ERP1", "Evet", "TEMELFATURA", "Tüzel Kişi",
                "ALTIN", "Websell",
            ]
        ],
    )

    sales_headers = [
        "MüşteriKodu", "FaturaNo", "Tarih", "KDV", "PersonelKodu", "ÖdemeTipi",
        "ÜrünKodu", "FOC", "Tabela Adı", "Vergi Dairesi", "Vergi No",
        "İlk Matbu No", "Fatura Kodu", "İrsaliye Kodu", "İrsaliye Numarası",
        "İrsaliye Tarihi", "Miktar", "Fiyat", "NetFiyat", "İskonto1",
        "EklenenKDV", "ToplamKDV", "Vade", "TuketiciFiyati", "İskonto2",
    ]
    sales = _save(
        tmp_path / "sales.xlsx",
        sales_headers,
        [[
            "1001", "F1", "20.07.2026", "D", "P1", "AH", "112064", "S",
            "TEST MARKET", "AYDIN", "1111111111", "F1", "FK", "IK", "I1",
            "20.07.2026", 20, 2.5, 59.7, 0, 0, 9.95, None, 3, 0.25,
        ]],
    )

    collections_headers = [
        "MusteriKodu", "Musteriİsmi", "BelgeNo", "BelgeTarihi", "TahsilatTipi",
        "TahsilatTuru", "SatisElemani", "Pesin/Diger", "Personel", "Rota",
        "MusteriKayitTipi", "MusteriTipi", "SahiplikTipi", "AltTip",
        "FiyatListesi", "BANKA", "Tutar",
    ]
    collections = _save(
        tmp_path / "collections.xlsx",
        collections_headers,
        [
            ["1001", "TEST MARKET", "B1", "20.07.2026", "N", "1", "P1", 0,
             "PERSONEL", "AYDIN-DD-0203", "Müşteri", "Müşteri",
             "Bagimsiz Satis Noktasi", "Market", "Liste", None, 100],
            ["1001", "TEST MARKET", "B2", "20.07.2026", "H", "1", None, 0,
             "", ".", "Müşteri", "Müşteri", "Bagimsiz Satis Noktasi",
             "Market", "Liste", "T.GARANTI BANKASI A.S.", 200],
        ],
    )

    result = ReportEditingEngine(
        [customer, sales, collections],
        resource_root=tmp_path,
        output_root=tmp_path / "out",
        create_template_outputs=False,
    ).run()

    assert result.customer_rows == 1
    assert result.sales_rows == 1
    assert result.collection_rows == 2
    assert result.collection_main_rows == 1
    assert result.output_dir is not None
    assert result.output_dir.name == "FOM AKTARMA - 20072026"

    customer_output = next(path for path in result.created_files if "MUSTERI" in path.name)
    customer_wb = load_workbook(customer_output, data_only=True)
    assert customer_wb.active["D2"].value == "SIMSEK-NAZILLI"

    cached_customer_list = refresh_customer_list_cache(result, tmp_path / "state")
    assert cached_customer_list is not None
    assert cached_customer_list.exists()
    cache_metadata = CustomerListCache(tmp_path / "state").metadata()
    assert cache_metadata is not None
    assert cache_metadata["orijinal_ad"] == customer.name

    sales_output = next(
        path for path in result.created_files
        if path.name.startswith(SALES_CLEAN_OUTPUT_PREFIX) and path.suffix == ".xlsx"
    )
    sales_wb = load_workbook(sales_output, data_only=True)
    sales_ws = sales_wb.active
    # 112064 urun kodu icin ozel Fiyat=NetFiyat/Miktar ve Iskonto2=0 kurali
    # kaldirildi; ham degerler oldugu gibi korunmali.
    assert sales_ws["R2"].value == 2.5
    assert sales_ws["T2"].value == 0.25
    assert sales_ws["W2"].value == 0
    assert sales_ws["Z2"].value == "SIMSEK-NAZILLI"

    collection_output = next(
        path for path in result.created_files
        if path.name.startswith(COLLECTION_CLEAN_OUTPUT_PREFIX) and path.suffix == ".xlsx"
    )
    collection_wb = load_workbook(collection_output, data_only=True)
    assert sales_wb.sheetnames == ["Sheet"]

    assert collection_wb.sheetnames == ["Sheet", "ŞUBELİLER"]
    assert collection_wb.worksheets[0].max_row == 2
    assert collection_wb.worksheets[0]["R2"].value == "SIMSEK-NAZILLI"
    assert collection_wb["ŞUBELİLER"].max_row == 3


def test_report_engine_accepts_sales_without_customer_list(tmp_path):
    sales = _save(
        tmp_path / "sales.xlsx",
        [
            "MüşteriKodu", "FaturaNo", "Tarih", "KDV", "PersonelKodu",
            "ÖdemeTipi", "ÜrünKodu", "FOC", "Tabela Adı", "Vergi Dairesi",
            "Vergi No", "İlk Matbu No", "Fatura Kodu", "İrsaliye Kodu",
            "İrsaliye Numarası", "İrsaliye Tarihi", "Miktar", "Fiyat",
            "NetFiyat", "İskonto1", "EklenenKDV", "ToplamKDV", "Vade",
            "TuketiciFiyati", "İskonto2",
        ],
        [[
            "C001", "F001", "22.07.2026", "D", "P01", "AH", "U01", "S",
            "TEST MÜŞTERİ", "MERKEZ", "1111111111", "M1", "F1", "I1", "IN1",
            "22.07.2026", 1, 100, 0, 0, 2, 18, 0, 100, 118,
        ]],
    )

    engine = ReportEditingEngine(
        [sales],
        resource_root=tmp_path,
        output_root=tmp_path / "out",
        create_template_outputs=False,
    )
    result = engine.run()

    assert result.sales_rows == 1
    assert result.customer_rows == 0
    assert any("tek başına düzenleniyor" in entry for entry in result.logs)
    sales_output = next(path for path in result.created_files if path.suffix == ".xlsx")
    assert load_workbook(sales_output, data_only=True).active["Z2"].value == "#N/A"


def test_report_engine_accepts_collections_without_customer_list(tmp_path):
    collections = _save(
        tmp_path / "collections.xlsx",
        [
            "MusteriKodu", "Musteriİsmi", "BelgeNo", "BelgeTarihi",
            "TahsilatTipi", "TahsilatTuru", "SatisElemani", "Pesin/Diger",
            "Personel", "Rota", "MusteriKayitTipi", "MusteriTipi",
            "SahiplikTipi", "AltTip", "FiyatListesi", "BANKA", "Tutar",
        ],
        [[
            "C001", "TEST MÜŞTERİ", "B001", "22.07.2026", "N", "1", "P01", 0,
            "PERSONEL", "AYDIN-DD-01", "Müşteri", "Müşteri", "Bağımsız", "Market",
            "Liste", "GARANTİ", 500,
        ]],
    )

    result = ReportEditingEngine(
        [collections],
        resource_root=tmp_path,
        output_root=tmp_path / "out",
        create_template_outputs=False,
    ).run()

    assert result.collection_rows == 1
    assert result.collection_main_rows == 1
    assert result.unmatched_customer_codes == 1
    assert any("tek başına düzenleniyor" in entry for entry in result.logs)
    collection_output = next(path for path in result.created_files if path.suffix == ".xlsx")
    workbook = load_workbook(collection_output, data_only=True)
    assert workbook.sheetnames == ["Sheet", "ŞUBELİLER"]
    assert workbook.active["R2"].value == "#N/A"


def test_original_template_outputs_keep_names_and_collection_contains_only_n1(tmp_path, monkeypatch):
    customer = _save(
        tmp_path / "customer.xlsx",
        _customer_headers(),
        [[
            "SIMSEK-BODRUM", 1, "1001", "TEST MARKET", "TEST LTD", "BODRUM",
            "1111111111", None, None, None, None, None, None, None, None,
            "Aktif", None, None, "MUĞLA", "BODRUM", 0, "Müşteri", "Müşteri",
            "Market", "Bagimsiz Satis Noktasi", "BODRUM-DD-0101", None,
            "Açık Hesap", None, 0, None, "555", 0, "Liste", "ERP1", "Evet",
            "TEMELFATURA", "Tüzel Kişi", "ALTIN", "Websell",
        ]],
    )
    sales = _save(
        tmp_path / "sales.xlsx",
        [
            "MüşteriKodu", "FaturaNo", "Tarih", "KDV", "PersonelKodu",
            "ÖdemeTipi", "ÜrünKodu", "FOC", "Tabela Adı", "Vergi Dairesi",
            "Vergi No", "İlk Matbu No", "Fatura Kodu", "İrsaliye Kodu",
            "İrsaliye Numarası", "İrsaliye Tarihi", "Miktar", "Fiyat",
            "NetFiyat", "İskonto1", "EklenenKDV", "ToplamKDV", "Vade",
            "TuketiciFiyati", "İskonto2",
        ],
        [[
            "1001", "F1", "21.07.2026", "D", "P1", "AH", "112064", "S",
            "TEST MARKET", "BODRUM", "1111111111", "F1", "FK", "IK", "I1",
            "21.07.2026", 2, 1000, 2101, 0, 0, 350, 0, 0, 9,
        ]],
    )
    collections = _save(
        tmp_path / "collections.xlsx",
        [
            "MusteriKodu", "Musteriİsmi", "BelgeNo", "BelgeTarihi",
            "TahsilatTipi", "TahsilatTuru", "SatisElemani", "Pesin/Diger",
            "Personel", "Rota", "MusteriKayitTipi", "MusteriTipi",
            "SahiplikTipi", "AltTip", "FiyatListesi", "BANKA", "Tutar",
        ],
        [
            ["1001", "TEST", "B1", "21.07.2026", "N", "1", "P1", 0, "PERSONEL",
             "BODRUM-DD-0101", "Müşteri", "Müşteri", "Bağımsız", "Market",
             "Liste", None, 100],
            ["1001", "TEST", "B2", "21.07.2026", "H", "1", "P1", 0, "PERSONEL",
             "BODRUM-DD-0101", "Müşteri", "Müşteri", "Bağımsız", "Market",
             "Liste", "BANKA", 200],
        ],
    )

    calls = []

    def fake_write(self, template_path, output_path, sheets, **kwargs):
        output_path = Path(output_path).with_suffix(".xls")
        output_path.write_bytes(b"test")
        calls.append((Path(output_path).name, sheets, kwargs))
        return output_path

    monkeypatch.setattr(ExcelTemplateWriter, "write", fake_write)

    result = ReportEditingEngine(
        [customer, sales, collections],
        resource_root=tmp_path,
        output_root=tmp_path / "out",
        create_template_outputs=True,
    ).run()

    names = {path.name for path in result.created_files}
    assert any(name.startswith(SALES_CLEAN_OUTPUT_PREFIX) and name.endswith(".xlsx") for name in names)
    assert f"{SALES_OUTPUT_BASENAME}.xls" in names
    assert any(name.startswith(COLLECTION_CLEAN_OUTPUT_PREFIX) and name.endswith(".xlsx") for name in names)
    assert f"{COLLECTION_OUTPUT_BASENAME}.xls" in names

    assert SALES_SHEET_NAME == "SATIS_FATURALARI"
    assert COLLECTION_SHEET_NAME == "TAHSILATLAR"
    assert len(SALES_SHEET_NAME + "$") <= 31
    assert len(COLLECTION_SHEET_NAME + "$") <= 31

    sales_call = next(call for call in calls if call[0] == f"{SALES_OUTPUT_BASENAME}.xls")
    assert sales_call[1][0][0] == SALES_SHEET_NAME
    assert sales_call[2]["delete_extra_sheets"] is True

    collection_call = next(
        call for call in calls if call[0] == f"{COLLECTION_OUTPUT_BASENAME}.xls"
    )
    assert len(collection_call[1]) == 1
    assert collection_call[1][0][0] == COLLECTION_SHEET_NAME
    assert len(collection_call[1][0][1]) == 1
    assert collection_call[1][0][1][0][4:6] == ["N", "1"]
    assert collection_call[2]["delete_extra_sheets"] is True


def test_classify_file_xls_satis_faturasi_taninir(tmp_path):
    """FOM artık satış faturası raporunu .xlsx yerine .xls (Excel 97-2003)
    olarak dışa aktarıyor; sütun yapısı aynı kalıyor. Regresyon testi."""
    path = tmp_path / "ENT-Muhasebe_Entegrasyon(Satış_Faturaları).xls"
    _save_xls(
        path,
        "cok_uzun_bir_sayfa_adi_31_karakteri_asan",
        _sales_headers(),
        [["C001", "F001", datetime(2026, 8, 4), "D", "P01", "AH", "U01", "S",
          "TEST MUSTERI", "MERKEZ", "1111111111", "M1", "F1", "I1", "IN1",
          datetime(2026, 8, 4), 1.0, 100.0, 0.0, 0.0, 18.0, 2.0, 0.0, 100.0, 118.0]],
    )
    assert ReportEditingEngine.classify_file(path) == "sales"


def test_read_rows_xls_tarih_datetime_olarak_gelir(tmp_path):
    """xlrd yoluyla okunan tarih hücreleri, openpyxl ile tutarlı olacak
    şekilde datetime nesnesi olarak gelmeli (ham Excel seri numarası değil)."""
    from app.modules.report_editing.engine import _read_rows

    path = tmp_path / "satis.xls"
    _save_xls(
        path,
        "SATIS",
        _sales_headers(),
        [["C001", "F001", datetime(2026, 8, 4), "D", "P01", "AH", "U01", "S",
          "TEST MUSTERI", "MERKEZ", "1111111111", "M1", "F1", "I1", "IN1",
          datetime(2026, 8, 4), 1.0, 100.0, 0.0, 0.0, 18.0, 2.0, 0.0, 100.0, 118.0]],
    )
    _, headers, rows = _read_rows(path)
    assert headers[:3] == ["MüşteriKodu", "FaturaNo", "Tarih"]
    assert isinstance(rows[0]["Tarih"], datetime)
    assert rows[0]["Tarih"].year == 2026
    assert rows[0]["Tarih"].month == 8
    assert rows[0]["Tarih"].day == 4


def test_xls_ve_xlsx_ayni_satis_verisini_ayni_sekilde_isler(tmp_path):
    """Aynı içerik .xlsx ve .xls olarak kaydedildiğinde motor aynı sayıda
    ve aynı içerikte satır üretmeli — format değişse de davranış aynı kalmalı."""
    headers = _sales_headers()
    row = ["C001", "F001", "04.08.2026", "D", "P01", "AH", "U01", "S",
           "TEST MUSTERI", "MERKEZ", "1111111111", "M1", "F1", "I1", "IN1",
           "04.08.2026", 1.0, 100.0, 0.0, 0.0, 18.0, 2.0, 0.0, 100.0, 118.0]

    xlsx_path = _save(tmp_path / "satis.xlsx", headers, [row])
    xls_path = _save_xls(tmp_path / "satis.xls", "SATIS", headers, [row])

    sheet_xlsx, rows_xlsx = ReportEditingEngine._sales_rows(xlsx_path, {})
    sheet_xls, rows_xls = ReportEditingEngine._sales_rows(xls_path, {})

    assert len(rows_xlsx) == len(rows_xls) == 1
    for key in headers:
        assert rows_xlsx[0][key] == rows_xls[0][key]
