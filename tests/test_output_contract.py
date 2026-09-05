from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
import xlwt

from app.core.output_contract import OutputContractError, validate_netsis_output
from app.core.output_profile import OutputProfileStore
from app.models.records import NetsisRecord


def _write_contract_file(path: Path, headers: list[str], bank_format: str) -> None:
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Sheet1")
    for column, header in enumerate(headers):
        sheet.write(0, column, header)
    bank_style = xlwt.easyxf(num_format_str=bank_format)
    sheet.write(1, 0, "SENTETIK-BANKA", bank_style)
    sheet.write(1, 14, 1250.50)
    workbook.save(str(path))


def test_toplu_output_contract_rejects_changed_bank_code_cell_format(tmp_path):
    profile = OutputProfileStore(Path(__file__).resolve().parents[1] / "config").get("netsis_toplu")
    template = tmp_path / "template.xls"
    output = tmp_path / "output.xls"
    _write_contract_file(template, profile.headers(), "0.00")
    _write_contract_file(output, profile.headers(), "@")
    records = [
        NetsisRecord(
            islem_tarihi=datetime(2026, 9, 5),
            cari_kodu="TEST001",
            tutar=1250.50,
            aciklama="TEST",
            banka="Garanti",
            bolge="BODRUM",
            kaynak="TEST",
            banka_hesap_kodu="SENTETIK-BANKA",
        )
    ]

    with pytest.raises(OutputContractError, match="hücre biçimi"):
        validate_netsis_output(output, profile, records, template)


def test_output_contract_rejects_amount_difference(tmp_path):
    profile = OutputProfileStore(Path(__file__).resolve().parents[1] / "config").get("netsis_toplu")
    output = tmp_path / "output.xls"
    _write_contract_file(output, profile.headers(), "0.00")
    records = [
        NetsisRecord(
            islem_tarihi=datetime(2026, 9, 5),
            cari_kodu="TEST001",
            tutar=1200,
            aciklama="TEST",
            banka="Garanti",
            bolge="BODRUM",
            kaynak="TEST",
            banka_hesap_kodu="SENTETIK-BANKA",
        )
    ]

    with pytest.raises(OutputContractError, match="toplamı uyuşmuyor"):
        validate_netsis_output(output, profile, records)


def test_xlsx_contract_preserves_template_sheet_set_and_two_bank_codes(tmp_path):
    profile = OutputProfileStore(Path(__file__).resolve().parents[1] / "config").get(
        "netsis_virman_toplu"
    )
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"

    for path in (template, output):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(profile.headers())
        sheet.append([
            "BANK-SOURCE", 1, None, "06.09.2026", "06.09.2026", None, None,
            "BANK-TARGET", None, None, None, None, 0, None, 1250, "TEST",
            None, None, "R00", 101, "TEST", None, None, None, None, None,
            None, None, None, None, None, None,
        ])
        workbook.create_sheet("Sayfa2")
        workbook.create_sheet("Sheet3")
        workbook.save(path)

    records = [
        type("Record", (), {"tutar": 1250})()
    ]
    validate_netsis_output(output, profile, records, template)


def test_xlsx_contract_rejects_removed_template_sheet(tmp_path):
    profile = OutputProfileStore(Path(__file__).resolve().parents[1] / "config").get(
        "netsis_virman_toplu"
    )
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    for path, extra_sheet in ((template, True), (output, False)):
        workbook = Workbook()
        workbook.active.append(profile.headers())
        if extra_sheet:
            workbook.create_sheet("Sayfa2")
        workbook.save(path)

    with pytest.raises(OutputContractError, match="sayfaları"):
        validate_netsis_output(output, profile, [], template)
