from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
import xlrd

from app.core.money import money_sum
from app.core.output_profile import OutputProfile
from app.models.records import NetsisRecord


class OutputContractError(ValueError):
    """Oluşan Excel dosyası seçili muhasebe profiliyle uyuşmadığında."""


def validate_netsis_output(
    output_path: str | Path,
    profile: OutputProfile,
    records: list[NetsisRecord],
    template_path: str | Path | None = None,
) -> None:
    """Netsis'e verilmeden önce başlık, satır, tutar ve kritik biçimi denetler."""
    output_path = Path(output_path)
    try:
        output = _WorkbookView.open(output_path)
    except Exception as error:
        raise OutputContractError(f"Çıktı Excel dosyası doğrulanamadı: {error}") from error

    expected_sheets = ["Sheet1"]
    template = None
    if template_path and Path(template_path).is_file():
        template = _WorkbookView.open(Path(template_path))
        expected_sheets = template.sheet_names
    if output.sheet_names != expected_sheets:
        raise OutputContractError(
            f"Netsis çıktı sayfaları onaylı şablonla uyuşmuyor: {output.sheet_names}"
        )
    expected_headers = profile.headers()
    actual_headers = [str(output.value(0, column) or "") for column in range(output.ncols)]
    if actual_headers != expected_headers:
        raise OutputContractError("Netsis çıktısının sütun başlıkları veya sırası değişmiş.")

    data_rows = [
        row
        for row in range(1, output.nrows)
        if any(str(output.value(row, column) or "").strip() for column in range(output.ncols))
    ]
    if len(data_rows) != len(records):
        raise OutputContractError(
            f"Netsis çıktı satır sayısı uyuşmuyor: beklenen {len(records)}, oluşan {len(data_rows)}."
        )

    amount_indexes = [
        index
        for index, column in enumerate(profile.columns)
        if column.source_kind == "field" and column.field == "tutar"
    ]
    if len(amount_indexes) != 1:
        raise OutputContractError("Netsis profilinde tam bir adet işlem tutarı sütunu bulunmalı.")
    amount_index = amount_indexes[0]
    try:
        output_total = money_sum(output.value(row, amount_index) for row in data_rows)
    except ValueError as error:
        raise OutputContractError(f"Netsis tutar sütununda sayısal olmayan değer var: {error}") from error
    expected_total = money_sum(record.tutar for record in records)
    if output_total != expected_total:
        raise OutputContractError(
            f"Netsis çıktı toplamı uyuşmuyor: beklenen {expected_total}, oluşan {output_total}."
        )

    bank_indexes = [
        index
        for index, column in enumerate(profile.columns)
        if column.source_kind == "field"
        and str(column.field or "").endswith("banka_hesap_kodu")
    ]
    if not bank_indexes:
        return
    missing_rows = [
        row + 1
        for row in data_rows
        if any(not str(output.value(row, index) or "").strip() for index in bank_indexes)
    ]
    if missing_rows:
        raise OutputContractError(f"Banka hesap kodu boş olan satırlar var: {missing_rows[:10]}")

    if template is None or not data_rows:
        return
    wrong_format_rows: list[int] = []
    for bank_index in bank_indexes:
        expected_format = template.number_format(1, bank_index)
        wrong_format_rows.extend(
            row + 1
            for row in data_rows
            if output.number_format(row, bank_index) != expected_format
        )
    if wrong_format_rows:
        raise OutputContractError(
            "Banka hesap kodu hücre biçimi onaylı şablondan farklı. "
            f"Kontrol edilmesi gereken satırlar: {sorted(set(wrong_format_rows))[:10]}"
        )


class _WorkbookView:
    def __init__(self, book, sheet, *, xlsx: bool):
        self.book = book
        self.sheet = sheet
        self.xlsx = xlsx
        self.sheet_names = list(book.sheetnames if xlsx else book.sheet_names())
        self.nrows = int(sheet.max_row if xlsx else sheet.nrows)
        self.ncols = int(sheet.max_column if xlsx else sheet.ncols)

    @classmethod
    def open(cls, path: Path) -> "_WorkbookView":
        if path.suffix.casefold() == ".xlsx":
            book = load_workbook(path, data_only=True, read_only=False)
            return cls(book, book.worksheets[0], xlsx=True)
        book = xlrd.open_workbook(str(path), formatting_info=True)
        return cls(book, book.sheet_by_index(0), xlsx=False)

    def value(self, row: int, column: int):
        if self.xlsx:
            return self.sheet.cell(row=row + 1, column=column + 1).value
        return self.sheet.cell_value(row, column)

    def number_format(self, row: int, column: int) -> str:
        if self.xlsx:
            return str(self.sheet.cell(row=row + 1, column=column + 1).number_format)
        xf = self.book.xf_list[self.sheet.cell_xf_index(row, column)]
        return self.book.format_map[xf.format_key].format_str
