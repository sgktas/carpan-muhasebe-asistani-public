from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.reconciliation_engine import ReconciliationResult


class ReconciliationReportWriter:
    HEADER_FILL = PatternFill(start_color="1F2B3A", end_color="1F2B3A", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    OK_FILL = PatternFill(start_color="DFF5E1", end_color="DFF5E1", fill_type="solid")
    WARN_FILL = PatternFill(start_color="FDEBEC", end_color="FDEBEC", fill_type="solid")

    def write(self, result: ReconciliationResult, output_path: str | Path) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        self._write_summary_sheet(workbook.active, result)
        self._write_diff_sheet(
            workbook.create_sheet("Sadece Bankada"),
            ["Tarih", "Açıklama", "Tutar", "Kaynak Dosya", "Kaynak Satır"],
            [
                [r.tarih, r.aciklama, r.tutar, r.kaynak_dosya, r.kaynak_satir]
                for r in result.sadece_bankada
            ],
        )
        self._write_diff_sheet(
            workbook.create_sheet("Sadece Netsis'te"),
            ["Tarih", "Açıklama", "Tutar", "Kaynak Dosya", "Kaynak Satır"],
            [
                [r.tarih, r.aciklama, r.tutar, r.kaynak_dosya, r.kaynak_satir]
                for r in result.sadece_netposte
            ],
        )

        workbook.save(output_path)
        return output_path

    def _write_summary_sheet(self, sheet, result: ReconciliationResult) -> None:
        sheet.title = "Özet"
        kalan_sayisi = len(result.sadece_bankada) + len(result.sadece_netposte)
        if result.mutabik and kalan_sayisi == 0:
            durum = "TAM MUTABIK"
        elif result.mutabik:
            durum = f"BAKİYE TUTUYOR — ama {kalan_sayisi} kayıt açıklanamadı, incelenmeli"
        else:
            durum = "MUTABIK DEĞİL — farkı inceleyin"

        rows = [
            ("Banka Bakiyesi (ay sonu)", result.banka_bakiyesi),
            ("Netsis Bakiyesi (ay sonu)", result.netsis_bakiyesi),
            ("Fark", result.fark),
            ("Eşleşen İşlem Sayısı", result.eslesen_sayisi),
            ("Bölünmüş Fiş Olarak Tanınan Grup Sayısı", result.bolunmus_grup_sayisi),
            ("Sadece Bankada Olan Sayısı", len(result.sadece_bankada)),
            ("Sadece Netsis'te Olan Sayısı", len(result.sadece_netposte)),
            ("Durum", durum),
        ]
        for row_index, (label, value) in enumerate(rows, start=1):
            sheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
            sheet.cell(row=row_index, column=2, value=value)
        for row_index in range(1, 4):
            sheet.cell(row=row_index, column=2).number_format = "#,##0.00"
        status_cell = sheet.cell(row=len(rows), column=2)
        status_cell.fill = self.OK_FILL if (result.mutabik and kalan_sayisi == 0) else self.WARN_FILL
        sheet.column_dimensions["A"].width = 42
        sheet.column_dimensions["B"].width = 52

    def _write_diff_sheet(self, sheet, headers: list[str], rows: list[list]) -> None:
        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_index, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        for row_index, row in enumerate(rows, start=2):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)
            sheet.cell(row=row_index, column=1).number_format = "dd.mm.yyyy hh:mm:ss"
            sheet.cell(row=row_index, column=3).number_format = "#,##0.00"
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 90
        sheet.column_dimensions["C"].width = 18
        sheet.column_dimensions["D"].width = 28
        sheet.column_dimensions["E"].width = 14
        sheet.freeze_panes = "A2"
