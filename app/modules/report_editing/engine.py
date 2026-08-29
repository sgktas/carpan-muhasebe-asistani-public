from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
import json
import os
from pathlib import Path
import re
import shutil
import subprocess
import tempfile
import unicodedata
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
import xlrd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CUSTOMER_OUTPUT_COLUMNS = [
    "Müşteri Sayısı", "Müşteri Kodu", "Şube", "Tabela Adi", "Ünvan",
    "Vergi Dairesi", "Vergi Numarası", "Tekel No", "SR-Rota", "Banka",
    "Ödeme Tipi", "Faks No", "Risk Limiti", "Etiket", "Telefon",
    "Kredi Limiti", "Fiyat Listesi", "ERP Kodu", "EFatura", "E-fat-Tip",
    "Vergi Tipi", "Hero Sınıfı", "Dış Kay.Sip.Kod",
]

SALES_OUTPUT_COLUMNS = [
    "MüşteriKodu", "FaturaNo", "Tarih", "KDV", "PersonelKodu", "ÖdemeTipi",
    "ÜrünKodu", "FOC", "Tabela Adı", "Vergi Dairesi", "Vergi No",
    "İlk Matbu No", "Fatura Kodu", "İrsaliye Kodu", "İrsaliye Numarası",
    "İrsaliye Tarihi", "Miktar", "Fiyat", "İskonto1", "İskonto2",
    "ToplamKDV", "EklenenKDV", "Vade", "TuketiciFiyati", "NetFiyat",
]

COLLECTION_OUTPUT_COLUMNS = [
    "MusteriKodu", "Musteriİsmi", "BelgeNo", "BelgeTarihi", "TahsilatTipi",
    "TahsilatTuru", "SatisElemani", "Pesin/Diger", "Personel", "Rota",
    "MusteriKayitTipi", "MusteriTipi", "SahiplikTipi", "AltTip",
    "FiyatListesi", "BANKA", "Tutar",
]


SALES_OUTPUT_BASENAME = "ENT-Muhasebe_Entegrasyon(Satış_Faturaları)"
COLLECTION_OUTPUT_BASENAME = "ENT-Muhasebe_Entegrasyon(Tahsilatlar)"
SALES_CLEAN_OUTPUT_PREFIX = "02_SATIS_RAPORU_DUZENLENMIS"
COLLECTION_CLEAN_OUTPUT_PREFIX = "03_TAHSILAT_RAPORU_DUZENLENMIS"
SALES_SHEET_NAME = "SATIS_FATURALARI"
COLLECTION_SHEET_NAME = "TAHSILATLAR"



def ensure_writable(path: str | Path) -> Path:
    path = Path(path)
    try:
        current_mode = path.stat().st_mode
        path.chmod(current_mode | 0o600)
    except OSError:
        pass
    if os.name == "nt":
        try:
            import ctypes
            attrs = ctypes.windll.kernel32.GetFileAttributesW(str(path))
            if attrs != 0xFFFFFFFF and attrs & 0x1:
                ctypes.windll.kernel32.SetFileAttributesW(str(path), attrs & ~0x1)
        except Exception:
            pass
    return path

MODULE_ID = "report_editing"
MODULE_NAME = "FOM Rapor Düzenleme"


@dataclass
class ReportEditingResult:
    output_dir: Path | None = None
    created_files: list[Path] = field(default_factory=list)
    logs: list[str] = field(default_factory=list)
    recognized_files: dict[str, Path] = field(default_factory=dict)
    customer_rows: int = 0
    sales_rows: int = 0
    collection_rows: int = 0
    collection_main_rows: int = 0
    unmatched_customer_codes: int = 0

    def summary(self) -> dict:
        return {
            "customer_rows": self.customer_rows,
            "sales_rows": self.sales_rows,
            "collection_rows": self.collection_rows,
            "collection_main_rows": self.collection_main_rows,
            "unmatched_customer_codes": self.unmatched_customer_codes,
            "created_file_count": len(self.created_files),
        }


def _normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^A-Z0-9]", "", text.upper().replace("İ", "I"))


def _normalize_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _clean_text(value: object):
    if isinstance(value, str):
        return value.replace("_x0009_", "\t")
    return value


def _safe_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    if not text:
        return None
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _date_values(rows: Sequence[dict], keys: Sequence[str]) -> list[datetime]:
    found: list[datetime] = []
    for row in rows:
        value = next((row.get(key) for key in keys if row.get(key) not in (None, "")), None)
        if value in (None, ""):
            continue
        if isinstance(value, datetime):
            found.append(value)
            continue
        text = str(value).strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                found.append(datetime.strptime(text[:10], fmt))
                break
            except ValueError:
                continue
    return found


def _format_date_label(dates: Sequence[datetime]) -> str:
    if not dates:
        return datetime.now().strftime("%d%m%Y")
    start = min(dates)
    end = max(dates)
    if start.date() == end.date():
        return start.strftime("%d%m%Y")
    if start.year == end.year and start.month == end.month:
        return f"{start:%d}-{end:%d.%m.%Y}"
    return f"{start:%d.%m.%Y}-{end:%d.%m.%Y}"


def _date_label(rows: Sequence[dict], keys: Sequence[str]) -> str:
    return _format_date_label(_date_values(rows, keys))


def _unique_output_dir(preferred: Path) -> Path:
    if not preferred.exists():
        return preferred
    suffix = 2
    while True:
        candidate = preferred.with_name(f"{preferred.name}_{suffix}")
        if not candidate.exists():
            return candidate
        suffix += 1


def _read_rows(path: Path) -> tuple[str, list[str], list[dict]]:
    if path.suffix.lower() == ".xls":
        return _read_rows_xls(path)

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        worksheet = workbook.worksheets[0]
        values = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = list(next(values))
        except StopIteration:
            raise ValueError(f"Excel dosyası boş: {path.name}")
        headers = [str(value).strip() if value is not None else "" for value in raw_headers]
        rows: list[dict] = []
        for values_row in values:
            row = {
                headers[index]: _clean_text(value)
                for index, value in enumerate(values_row[: len(headers)])
                if headers[index]
            }
            if any(value not in (None, "") for value in row.values()):
                rows.append(row)
        return worksheet.title, headers, rows
    finally:
        workbook.close()


def _read_rows_xls(path: Path) -> tuple[str, list[str], list[dict]]:
    """FOM'un artık Excel 97-2003 (.xls) olarak dışa aktardığı dosyalar için
    ``_read_rows`` ile birebir aynı sözleşmeyi (sayfa adı, başlıklar, satırlar)
    xlrd ile üretir. Sütun yapısı openpyxl yoluyla aynıdır; yalnızca dosya
    biçimi farklıdır. Tarih hücreleri, openpyxl'in davranışıyla tutarlı olacak
    şekilde ``datetime`` nesnesine çevrilir — aksi halde xlrd bunları ham
    Excel seri numarası (örn. 45870.0) olarak döndürür.
    """
    workbook = xlrd.open_workbook(str(path))
    worksheet = workbook.sheet_by_index(0)
    if worksheet.nrows == 0:
        raise ValueError(f"Excel dosyası boş: {path.name}")

    def cell_python_value(row_index: int, col_index: int):
        cell = worksheet.cell(row_index, col_index)
        if cell.ctype == xlrd.XL_CELL_DATE:
            return xlrd.xldate_as_datetime(cell.value, workbook.datemode)
        return cell.value

    raw_headers = [cell_python_value(0, col) for col in range(worksheet.ncols)]
    headers = [str(value).strip() if value not in (None, "") else "" for value in raw_headers]

    rows: list[dict] = []
    for row_index in range(1, worksheet.nrows):
        row = {
            headers[col]: _clean_text(cell_python_value(row_index, col))
            for col in range(min(worksheet.ncols, len(headers)))
            if headers[col]
        }
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return worksheet.name, headers, rows


def _write_clean_xlsx(
    output_path: Path,
    sheet_payloads: Sequence[tuple[str, list[str], list[list[object]], bool]],
    *,
    blank_first_column: bool = False,
    number_formats_by_sheet: Sequence[dict[str, str]] | None = None,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_fill = PatternFill("solid", fgColor="214866")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="DDE4EA")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_index, (sheet_name, headers, rows, use_filter) in enumerate(sheet_payloads):
        worksheet = workbook.create_sheet(title=(sheet_name or f"Rapor{sheet_index + 1}")[:31])
        start_column = 2 if blank_first_column else 1

        for col_offset, header in enumerate(headers, start=start_column):
            cell = worksheet.cell(1, col_offset, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border

        max_lengths = [len(str(header or "")) for header in headers]
        for row_index, row in enumerate(rows, start=2):
            for col_offset, value in enumerate(row, start=start_column):
                worksheet.cell(row_index, col_offset, value)
                logical_index = col_offset - start_column
                if logical_index < len(max_lengths):
                    max_lengths[logical_index] = min(
                        70, max(max_lengths[logical_index], len(str(value or "")))
                    )

        for logical_index, max_length in enumerate(max_lengths):
            column_index = start_column + logical_index
            worksheet.column_dimensions[get_column_letter(column_index)].width = max(
                10, min(max_length + 2, 52)
            )

        sheet_number_formats = (
            number_formats_by_sheet[sheet_index]
            if number_formats_by_sheet and sheet_index < len(number_formats_by_sheet)
            else {}
        )
        for header_name, number_format in sheet_number_formats.items():
            if header_name not in headers:
                continue
            column_index = start_column + headers.index(header_name)
            for row_index in range(2, len(rows) + 2):
                worksheet.cell(row_index, column_index).number_format = number_format

        worksheet.freeze_panes = f"{get_column_letter(start_column)}2"
        if use_filter and rows:
            end_column = get_column_letter(start_column + len(headers) - 1)
            worksheet.auto_filter.ref = (
                f"{get_column_letter(start_column)}1:{end_column}{len(rows) + 1}"
            )

    workbook.save(output_path)
    return ensure_writable(output_path)


def _report_template_path(resource_root: Path, file_name: str) -> Path:
    local = resource_root / "templates" / "local" / "report_editing" / file_name
    if os.environ.get("MUHASEBE_ASISTANI_DISABLE_LOCAL_CONFIG") != "1" and local.is_file():
        return local
    return resource_root / "templates" / "report_editing" / file_name


class ExcelTemplateWriter:
    """Veriyi kullanıcının gerçek Excel 97-2003 şablonlarına yazar."""

    def write(
        self,
        template_path: str | Path,
        output_path: str | Path,
        sheets: Sequence[tuple[str | None, list[list[object]]]],
        *,
        headers: Sequence[list[str]] | None = None,
        delete_extra_sheets: bool = False,
        number_formats: dict[tuple[int, int], tuple[str, str]] | None = None,
    ) -> Path:
        template_path = Path(template_path)
        output_path = Path(output_path).with_suffix(".xls")
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if not template_path.is_file():
            return self._write_xlwt(output_path, sheets, headers=headers)

        if os.name != "nt":
            return self._write_xlwt(output_path, sheets, headers=headers)

        try:
            return self._write_pywin32(
                template_path,
                output_path,
                sheets,
                delete_extra_sheets=delete_extra_sheets,
                number_formats=number_formats,
            )
        except ModuleNotFoundError:
            return self._write_powershell(
                template_path,
                output_path,
                sheets,
                delete_extra_sheets=delete_extra_sheets,
                number_formats=number_formats,
            )

    @staticmethod
    def _write_xlwt(
        output_path: Path,
        sheets: Sequence[tuple[str | None, list[list[object]]]],
        *,
        headers: Sequence[list[str]] | None,
    ) -> Path:
        import xlwt

        from app.writers.xls_utils import make_styles, safe_sheet_name, save_xls, write_cell

        workbook = xlwt.Workbook(encoding="utf-8", style_compression=2)
        styles = make_styles()
        used_names: set[str] = set()
        amount_headers = {
            "Tutar", "Fiyat", "Miktar", "İskonto1", "İskonto2",
            "ToplamKDV", "EklenenKDV", "TuketiciFiyati", "NetFiyat",
        }
        date_headers = {"Tarih", "İrsaliye Tarihi", "BelgeTarihi"}

        for sheet_index, (desired_name, values) in enumerate(sheets):
            sheet_headers = list(headers[sheet_index]) if headers and sheet_index < len(headers) else []
            column_count = max(len(sheet_headers), max((len(row) for row in values), default=0))
            if len(values) + 1 > 65_536:
                raise ValueError("Excel 97-2003 en fazla 65.536 satır destekler.")
            worksheet = workbook.add_sheet(safe_sheet_name(desired_name or "Sheet1", used_names))
            worksheet.panes_frozen = True
            worksheet.horz_split_pos = 1

            for column_index in range(column_count):
                header = sheet_headers[column_index] if column_index < len(sheet_headers) else ""
                worksheet.write(0, column_index, header, styles["header"])
                worksheet.col(column_index).width = min(max(12, len(str(header)) + 2) * 256, 65_535)

            for row_index, row in enumerate(values, start=1):
                for column_index in range(column_count):
                    value = row[column_index] if column_index < len(row) else ""
                    header = sheet_headers[column_index] if column_index < len(sheet_headers) else ""
                    style_name = "amount" if header in amount_headers else "date" if header in date_headers else None
                    write_cell(worksheet, row_index, column_index, value, styles, style_name)

        return save_xls(workbook, output_path)

    @staticmethod
    def _python_matrix(values: list[list[object]]) -> tuple[tuple[object, ...], ...]:
        return tuple(
            tuple(
                value.strftime("%d.%m.%Y") if isinstance(value, datetime) else value
                for value in row
            )
            for row in values
        )

    def _write_pywin32(
        self,
        template_path: Path,
        output_path: Path,
        sheets: Sequence[tuple[str | None, list[list[object]]]],
        *,
        delete_extra_sheets: bool,
        number_formats: dict[tuple[int, int], tuple[str, str]] | None,
    ) -> Path:
        import pythoncom
        import win32com.client

        pythoncom.CoInitialize()
        excel = None
        workbook = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.EnableEvents = False
            excel.ScreenUpdating = False
            workbook = excel.Workbooks.Open(str(template_path.resolve()), 0, False)
            workbook.CheckCompatibility = False

            for index, (desired_name, values) in enumerate(sheets, start=1):
                if index <= workbook.Worksheets.Count:
                    worksheet = workbook.Worksheets(index)
                else:
                    workbook.Worksheets(1).Copy(
                        After=workbook.Worksheets(workbook.Worksheets.Count)
                    )
                    worksheet = workbook.Worksheets(workbook.Worksheets.Count)

                if desired_name:
                    try:
                        worksheet.Name = desired_name[:31]
                    except Exception:
                        pass

                row_count = len(values)
                column_count = max((len(row) for row in values), default=1)
                used_rows = int(worksheet.UsedRange.Rows.Count)
                used_columns = int(worksheet.UsedRange.Columns.Count)
                clear_last = max(used_rows, row_count + 1, 2)
                worksheet.Range(
                    worksheet.Cells(2, 1),
                    worksheet.Cells(clear_last, max(column_count, used_columns)),
                ).ClearContents()

                if row_count:
                    worksheet.Range(
                        worksheet.Cells(2, 1),
                        worksheet.Cells(row_count + 1, column_count),
                    ).Value2 = self._python_matrix(values)

                for (format_sheet_index, column_index), formats in (number_formats or {}).items():
                    if format_sheet_index != index or row_count <= 0:
                        continue
                    invariant_format, local_format = formats
                    target_range = worksheet.Range(
                        worksheet.Cells(2, column_index),
                        worksheet.Cells(row_count + 1, column_index),
                    )
                    try:
                        target_range.NumberFormatLocal = local_format
                    except Exception:
                        target_range.NumberFormat = invariant_format

            if delete_extra_sheets:
                while workbook.Worksheets.Count > len(sheets):
                    workbook.Worksheets(workbook.Worksheets.Count).Delete()

            output_path.unlink(missing_ok=True)
            workbook.SaveAs(str(output_path.resolve()), FileFormat=56)
            workbook.Close(False)
            workbook = None
        finally:
            if workbook is not None:
                try:
                    workbook.Close(False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

        if not output_path.is_file():
            raise OSError(f"Şablon çıktısı oluşturulamadı: {output_path}")
        return ensure_writable(output_path)

    def _write_powershell(
        self,
        template_path: Path,
        output_path: Path,
        sheets: Sequence[tuple[str | None, list[list[object]]]],
        *,
        delete_extra_sheets: bool,
        number_formats: dict[tuple[int, int], tuple[str, str]] | None,
    ) -> Path:
        powershell = shutil.which("powershell.exe") or shutil.which("powershell")
        if not powershell:
            raise RuntimeError(
                "Microsoft Excel şablon çıktısı için pywin32 veya Windows PowerShell bulunamadı."
            )

        payload = {
            "delete_extra_sheets": delete_extra_sheets,
            "number_formats": [
                {
                    "sheet_index": sheet_index,
                    "column_index": column_index,
                    "invariant": formats[0],
                    "local": formats[1],
                }
                for (sheet_index, column_index), formats in (number_formats or {}).items()
            ],
            "sheets": [
                {
                    "name": name,
                    "rows": [
                        [
                            value.strftime("%d.%m.%Y") if isinstance(value, datetime) else value
                            for value in row
                        ]
                        for row in rows
                    ],
                }
                for name, rows in sheets
            ]
        }

        script = r"""param(
    [Parameter(Mandatory = $true)][string]$TemplatePath,
    [Parameter(Mandatory = $true)][string]$OutputPath,
    [Parameter(Mandatory = $true)][string]$DataPath
)
$ErrorActionPreference = "Stop"
$excel = $null
$workbook = $null
try {
    $payload = Get-Content -LiteralPath $DataPath -Raw -Encoding UTF8 | ConvertFrom-Json
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $excel.EnableEvents = $false
    $excel.ScreenUpdating = $false
    $workbook = $excel.Workbooks.Open($TemplatePath, 0, $false)
    $workbook.CheckCompatibility = $false

    $sheetPayloads = @($payload.sheets)
    for ($sheetIndex = 0; $sheetIndex -lt $sheetPayloads.Count; $sheetIndex++) {
        if (($sheetIndex + 1) -le $workbook.Worksheets.Count) {
            $worksheet = $workbook.Worksheets.Item($sheetIndex + 1)
        }
        else {
            $workbook.Worksheets.Item(1).Copy(
                [System.Type]::Missing,
                $workbook.Worksheets.Item($workbook.Worksheets.Count)
            )
            $worksheet = $workbook.Worksheets.Item($workbook.Worksheets.Count)
        }

        $sheetPayload = $sheetPayloads[$sheetIndex]
        if ($null -ne $sheetPayload.name -and [string]$sheetPayload.name -ne "") {
            try {
                $nameText = [string]$sheetPayload.name
                $worksheet.Name = $nameText.Substring(0, [Math]::Min(31, $nameText.Length))
            } catch {}
        }

        $rows = @($sheetPayload.rows)
        $columnCount = 1
        foreach ($row in $rows) {
            $rowArray = @($row)
            if ($rowArray.Count -gt $columnCount) { $columnCount = $rowArray.Count }
        }

        $usedRows = [int]$worksheet.UsedRange.Rows.Count
        $usedColumns = [int]$worksheet.UsedRange.Columns.Count
        $clearLast = [Math]::Max([Math]::Max($usedRows, $rows.Count + 1), 2)
        $clearColumns = [Math]::Max($usedColumns, $columnCount)
        $worksheet.Range(
            $worksheet.Cells.Item(2, 1),
            $worksheet.Cells.Item($clearLast, $clearColumns)
        ).ClearContents()

        if ($rows.Count -gt 0) {
            $matrix = [System.Array]::CreateInstance(
                [object], [int[]]@($rows.Count, $columnCount)
            )
            for ($rowIndex = 0; $rowIndex -lt $rows.Count; $rowIndex++) {
                $values = @($rows[$rowIndex])
                for ($columnIndex = 0; $columnIndex -lt $columnCount; $columnIndex++) {
                    $value = $null
                    if ($columnIndex -lt $values.Count) { $value = $values[$columnIndex] }
                    $matrix.SetValue($value, $rowIndex, $columnIndex)
                }
            }
            $worksheet.Range(
                $worksheet.Cells.Item(2, 1),
                $worksheet.Cells.Item($rows.Count + 1, $columnCount)
            ).Value2 = $matrix
        }
        $numberFormats = @($payload.number_formats)
        foreach ($formatRule in $numberFormats) {
            if ([int]$formatRule.sheet_index -ne ($sheetIndex + 1) -or $rows.Count -le 0) {
                continue
            }
            $formatRange = $worksheet.Range(
                $worksheet.Cells.Item(2, [int]$formatRule.column_index),
                $worksheet.Cells.Item($rows.Count + 1, [int]$formatRule.column_index)
            )
            try {
                $formatRange.NumberFormatLocal = [string]$formatRule.local
            }
            catch {
                $formatRange.NumberFormat = [string]$formatRule.invariant
            }
            [void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($formatRange)
        }
        [void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($worksheet)
    }

    if ([bool]$payload.delete_extra_sheets) {
        while ($workbook.Worksheets.Count -gt $sheetPayloads.Count) {
            $workbook.Worksheets.Item($workbook.Worksheets.Count).Delete()
        }
    }

    if (Test-Path -LiteralPath $OutputPath) {
        Remove-Item -LiteralPath $OutputPath -Force
    }
    $workbook.SaveAs($OutputPath, 56)
    $workbook.Close($false)
    $workbook = $null
}
finally {
    if ($workbook -ne $null) { try { $workbook.Close($false) } catch {} }
    if ($excel -ne $null) { try { $excel.Quit() } catch {} }
    if ($workbook -ne $null) {
        [void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($workbook)
    }
    if ($excel -ne $null) {
        [void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($excel)
    }
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}
"""

        with tempfile.TemporaryDirectory(prefix="carpan_report_template_") as temp_dir:
            temp_dir_path = Path(temp_dir)
            data_path = temp_dir_path / "payload.json"
            script_path = temp_dir_path / "write_template.ps1"
            data_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8-sig")
            script_path.write_text(script, encoding="utf-8-sig")
            completed = subprocess.run(
                [
                    powershell,
                    "-NoLogo", "-NoProfile", "-NonInteractive",
                    "-ExecutionPolicy", "Bypass",
                    "-File", str(script_path),
                    "-TemplatePath", str(template_path.resolve()),
                    "-OutputPath", str(output_path.resolve()),
                    "-DataPath", str(data_path.resolve()),
                ],
                capture_output=True,
                text=True,
                errors="replace",
                check=False,
                creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
            )
            if completed.returncode != 0:
                details = (completed.stderr or completed.stdout or "Bilinmeyen hata").strip()
                raise RuntimeError(f"Orijinal Excel şablonu doldurulamadı: {details}")

        if not output_path.is_file():
            raise OSError(f"Şablon çıktısı oluşturulamadı: {output_path}")
        return ensure_writable(output_path)


class ReportEditingEngine:
    def __init__(
        self,
        files: Iterable[str | Path],
        *,
        resource_root: str | Path,
        output_root: str | Path,
        create_template_outputs: bool = True,
    ):
        self.files = [Path(path) for path in files]
        self.resource_root = Path(resource_root)
        self.output_root = Path(output_root)
        self.create_template_outputs = create_template_outputs

    @staticmethod
    def classify_file(path: str | Path) -> str | None:
        path = Path(path)
        suffix = path.suffix.lower()
        if suffix not in (".xlsx", ".xls"):
            return None

        if suffix == ".xls":
            workbook = xlrd.open_workbook(str(path))
            worksheet = workbook.sheet_by_index(0)
            if worksheet.nrows == 0:
                return None
            headers = [_normalize_header(value) for value in worksheet.row_values(0)]
        else:
            workbook = load_workbook(path, read_only=True, data_only=False)
            try:
                headers = [_normalize_header(cell.value) for cell in workbook.worksheets[0][1]]
            finally:
                workbook.close()

        header_set = set(headers)
        if {"SUBE", "MUSTERISAYISI", "MUSTERIKODU", "VERGINUMARASI"}.issubset(header_set):
            return "customer"
        if {"MUSTERIKODU", "FATURANO", "URUNKODU", "NETFIYAT"}.issubset(header_set):
            return "sales"
        if {"MUSTERIKODU", "BELGENO", "TAHSILATTIPI", "TUTAR"}.issubset(header_set):
            return "collections"
        return None

    def _classify(self) -> dict[str, Path]:
        recognized: dict[str, Path] = {}
        unknown: list[str] = []
        for path in self.files:
            file_type = self.classify_file(path)
            if file_type is None:
                unknown.append(path.name)
                continue
            if file_type in recognized:
                raise ValueError(
                    f"Aynı rapor türünden birden fazla dosya seçildi: "
                    f"{path.name} / {recognized[file_type].name}"
                )
            recognized[file_type] = path
        if unknown:
            raise ValueError("Rapor türü tanınamayan dosya(lar): " + ", ".join(unknown))
        if not recognized:
            raise ValueError("Düzenlenecek müşteri, satış veya tahsilat raporu bulunamadı.")
        if ("sales" in recognized or "collections" in recognized) and "customer" not in recognized:
            raise ValueError(
                "Satış veya tahsilat raporuna şube bilgisi eklemek için ham müşteri listesi de seçilmelidir."
            )
        return recognized

    @staticmethod
    def _customer_rows(path: Path) -> tuple[str, list[dict], dict[str, str]]:
        sheet_name, headers, rows = _read_rows(path)
        missing = [header for header in CUSTOMER_OUTPUT_COLUMNS if header not in headers]
        if missing:
            raise ValueError("Müşteri listesinde gerekli sütunlar eksik: " + ", ".join(missing))

        cleaned: list[dict] = []
        branch_lookup: dict[str, str] = {}
        for source in rows:
            row = {column: source.get(column) for column in CUSTOMER_OUTPUT_COLUMNS}
            route = str(row.get("SR-Rota") or "").upper()
            branch = str(row.get("Şube") or "").upper()
            if branch == "SIMSEK-AYDIN" and (
                "AYDIN-DD-02" in route or "AYDIN-WHS-02" in route
            ):
                row["Şube"] = "SIMSEK-NAZILLI"
            code = _normalize_code(row.get("Müşteri Kodu"))
            if code:
                branch_lookup[code] = str(row.get("Şube") or "")
            cleaned.append(row)
        return sheet_name, cleaned, branch_lookup

    @staticmethod
    def _sales_rows(path: Path, branch_lookup: dict[str, str]) -> tuple[str, list[dict]]:
        sheet_name, headers, rows = _read_rows(path)
        missing = [header for header in SALES_OUTPUT_COLUMNS if header not in headers]
        if missing:
            raise ValueError("Satış raporunda gerekli sütunlar eksik: " + ", ".join(missing))

        cleaned: list[dict] = []
        for source in rows:
            row = {column: source.get(column) for column in SALES_OUTPUT_COLUMNS}
            if row.get("Vade") in (None, ""):
                row["Vade"] = 0
            row["Şube"] = branch_lookup.get(
                _normalize_code(row.get("MüşteriKodu")), "#N/A"
            )
            cleaned.append(row)
        return sheet_name, cleaned

    @staticmethod
    def _collection_rows(
        path: Path, branch_lookup: dict[str, str]
    ) -> tuple[str, list[dict], list[dict], int]:
        sheet_name, headers, rows = _read_rows(path)
        missing = [header for header in COLLECTION_OUTPUT_COLUMNS if header not in headers]
        if missing:
            raise ValueError("Tahsilat raporunda gerekli sütunlar eksik: " + ", ".join(missing))

        all_rows: list[dict] = []
        main_rows: list[dict] = []
        unmatched = 0
        for source in rows:
            row = {column: source.get(column) for column in COLLECTION_OUTPUT_COLUMNS}
            code = _normalize_code(row.get("MusteriKodu"))
            branch = branch_lookup.get(code) if code else None
            if not branch:
                branch = "#N/A"
                if code:
                    unmatched += 1
            row["Şube"] = branch
            all_rows.append(row)
            if (
                str(row.get("TahsilatTipi") or "").strip().upper() == "N"
                and str(row.get("TahsilatTuru") or "").strip() == "1"
            ):
                main_rows.append(dict(row))
        return sheet_name, main_rows, all_rows, unmatched

    def run(self) -> ReportEditingResult:
        result = ReportEditingResult()
        recognized = self._classify()
        result.recognized_files = dict(recognized)

        operation_dates: list[datetime] = []
        if "sales" in recognized:
            _, _, source_rows = _read_rows(recognized["sales"])
            operation_dates.extend(_date_values(source_rows, ("Tarih", "İrsaliye Tarihi")))
        if "collections" in recognized:
            _, _, source_rows = _read_rows(recognized["collections"])
            operation_dates.extend(_date_values(source_rows, ("BelgeTarihi",)))

        operation_date_label = _format_date_label(operation_dates)
        output_dir = _unique_output_dir(
            self.output_root / f"FOM AKTARMA - {operation_date_label}"
        )
        output_dir.mkdir(parents=True, exist_ok=False)
        result.output_dir = output_dir
        result.logs.append(f"FOM işlem tarihi: {operation_date_label}")

        customer_rows: list[dict] = []
        branch_lookup: dict[str, str] = {}
        if "customer" in recognized:
            customer_sheet, customer_rows, branch_lookup = self._customer_rows(
                recognized["customer"]
            )
            customer_values = [
                [row.get(column) for column in CUSTOMER_OUTPUT_COLUMNS]
                for row in customer_rows
            ]
            output = output_dir / "01_MUSTERI_LISTESI_DUZENLENMIS.xlsx"
            _write_clean_xlsx(
                output,
                [(customer_sheet, CUSTOMER_OUTPUT_COLUMNS, customer_values, False)],
                blank_first_column=True,
            )
            result.created_files.append(output)
            result.customer_rows = len(customer_rows)
            result.logs.append(
                f"Müşteri listesi düzenlendi: {len(customer_rows)} kayıt. "
                "AYDIN-DD-02 rotaları SIMSEK-NAZILLI şubesine ayrıldı."
            )

        sales_rows: list[dict] = []
        if "sales" in recognized:
            sales_sheet, sales_rows = self._sales_rows(recognized["sales"], branch_lookup)
            headers = SALES_OUTPUT_COLUMNS + [""]
            values = [
                [row.get(column) for column in SALES_OUTPUT_COLUMNS] + [row.get("Şube")]
                for row in sales_rows
            ]
            label = _date_label(sales_rows, ("Tarih", "İrsaliye Tarihi"))
            clean_output = output_dir / f"{SALES_CLEAN_OUTPUT_PREFIX}_{label}.xlsx"
            _write_clean_xlsx(
                clean_output,
                [(sales_sheet, headers, values, False)],
            )
            result.created_files.append(clean_output)
            result.sales_rows = len(sales_rows)
            result.logs.append(
                f"Satış raporu düzenlendi: {len(sales_rows)} kayıt. Boş vadeler 0 yapıldı."
            )

            if self.create_template_outputs:
                template_output = output_dir / f"{SALES_OUTPUT_BASENAME}.xls"
                ExcelTemplateWriter().write(
                    _report_template_path(self.resource_root, "sales_template.xls"),
                    template_output,
                    [(SALES_SHEET_NAME, values)],
                    headers=[SALES_OUTPUT_COLUMNS + [""]],
                    delete_extra_sheets=True,
                )
                result.created_files.append(template_output)
                result.logs.append("Satış verileri orijinal Excel 97-2003 şablonuna yazıldı.")

        if "collections" in recognized:
            sheet_name, main_rows, all_rows, unmatched = self._collection_rows(
                recognized["collections"], branch_lookup
            )
            headers = COLLECTION_OUTPUT_COLUMNS + [""]
            main_values = [
                [row.get(column) for column in COLLECTION_OUTPUT_COLUMNS] + [row.get("Şube")]
                for row in main_rows
            ]
            all_values = [
                [row.get(column) for column in COLLECTION_OUTPUT_COLUMNS] + [row.get("Şube")]
                for row in all_rows
            ]
            label = _date_label(all_rows, ("BelgeTarihi",))
            clean_output = output_dir / f"{COLLECTION_CLEAN_OUTPUT_PREFIX}_{label}.xlsx"
            _write_clean_xlsx(
                clean_output,
                [
                    (sheet_name, headers, main_values, False),
                    ("ŞUBELİLER", headers, all_values, True),
                ],
            )
            result.created_files.append(clean_output)
            result.collection_rows = len(all_rows)
            result.collection_main_rows = len(main_rows)
            result.unmatched_customer_codes = unmatched
            result.logs.append(
                f"Tahsilat raporu düzenlendi: {len(all_rows)} toplam kayıt; "
                f"{len(main_rows)} N/1 kaydı ana sayfaya alındı."
            )
            if unmatched:
                result.logs.append(
                    f"UYARI: {unmatched} tahsilat satırında müşteri kodu bulunamadı; "
                    "Şube '#N/A' bırakıldı."
                )

            if self.create_template_outputs:
                template_output = output_dir / f"{COLLECTION_OUTPUT_BASENAME}.xls"
                ExcelTemplateWriter().write(
                    _report_template_path(self.resource_root, "collections_template.xls"),
                    template_output,
                    [(COLLECTION_SHEET_NAME, main_values)],
                    headers=[COLLECTION_OUTPUT_COLUMNS + ["BÖLGE"]],
                    delete_extra_sheets=True,
                )
                result.created_files.append(template_output)
                result.logs.append("Tahsilat verileri orijinal Excel 97-2003 şablonuna yazıldı.")

        if not result.created_files:
            raise RuntimeError("Rapor düzenleme işlemi çıktı üretemedi.")
        return result
